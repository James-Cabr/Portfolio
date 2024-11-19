import os
import arcpy
import smartsheet
import requests
from datetime import datetime
import pytz
import concurrent.futures
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Initialize the Smartsheet client using the API token.
# Replace the placeholder with your actual API token.
API_TOKEN = 'REPLACE_WITH_SMARTSHEET_API_TOKEN'
smartsheet_client = smartsheet.Smartsheet(API_TOKEN)

# Define key fields and feature service URL.
# Replace these placeholders with your actual feature service URL and file paths.
row_id_field = 'Row_ID'
feature_service_url = r"https://utility.arcgis.com/usrsvcs/servers/REPLACE_WITH_FEATURE_SERVICE_URL/rest/services/PROJECT_NAME/FeatureServer/1"
download_folder = r"C:\REPLACE_WITH_DOWNLOAD_FOLDER_PATH"

# Define how fields in Smartsheet map to fields in the feature service.
field_mapping = {
    'Row_ID': 'Row_ID', 'SMARTNAME': 'SMARTNAME', 'Parcel ID Number(s)': 'Parcel_ID_Numbers'
    # Add the rest of the field mappings here as needed.
}

# Helper function to format numeric values as strings.
def format_value(value):
    """Ensure numeric values are returned as formatted strings to prevent issues."""
    if isinstance(value, float):
        return "{:.10f}".format(value).rstrip('0').rstrip('.')
    return str(value)

# Helper function to handle date conversion to a readable format in a specific timezone.
def format_date(date_str, timezone_str='America/Los_Angeles'):
    """Convert an ISO 8601 date string to a readable local time format."""
    try:
        date_obj = datetime.strptime(date_str[:-1], "%Y-%m-%dT%H:%M:%S")
        date_obj = pytz.utc.localize(date_obj)
        local_tz = pytz.timezone(timezone_str)
        local_date = date_obj.astimezone(local_tz)
        return local_date.strftime("%m/%d/%y %I:%M %p")
    except ValueError:
        # If the input isn't a valid ISO 8601 string, just return it as-is.
        return date_str

# Function to update the feature service with data from Smartsheet.
def append_to_feature_service(fc, smartsheet_data, field_mapping, parcel_id_list):
    """Update feature service records with Smartsheet data."""
    update_fields = list(field_mapping.values())
    parcel_id_field = 'Parcel_ID'
    smartsheet_parcel_id_field = 'Parcel_ID_Numbers'

    # Identify and remove any Parcel IDs that are actually geographic coordinates.
    coordinate_pattern = re.compile(r"\d{1,3}°\s?\d{1,2}'\d{1,2}\.\d{1,2}\"[NS]?,\s?-?\s?\d{1,3}°\s?\d{1,2}'\d{1,2}\.\d{1,2}\"[EW]?")
    filtered_parcel_ids = [pid for pid in parcel_id_list if not coordinate_pattern.match(pid)]

    # Construct a SQL query to select rows that match the filtered Parcel IDs.
    sql_query = f"{parcel_id_field} IN ('{', '.join(filtered_parcel_ids)}')"
    print(sql_query)

    # Update rows in the feature service where the Parcel_ID matches.
    with arcpy.da.UpdateCursor(fc, [parcel_id_field] + update_fields, where_clause=sql_query) as cursor:
        for row in cursor:
            parcel_id = row[0]  # First column is Parcel_ID.
            if parcel_id in parcel_id_list:
                # Find the corresponding record in Smartsheet data.
                record = next((rec for rec in smartsheet_data.values() if rec.get(smartsheet_parcel_id_field) == parcel_id), None)
                if record:
                    # Update each field with the value from the corresponding Smartsheet record.
                    for i, field in enumerate(update_fields, start=1):
                        row[i] = record.get(field_mapping.get(field, field), None)
                    cursor.updateRow(row)
                    print(f"Updated Parcel_ID {parcel_id} with new data.")

# Fetch all Smartsheet sheets and identify the one to process.
response = smartsheet_client.Sheets.list_sheets(include_all=True)
sheets = response.data
parcel_id_list = []

sheet_id = next((sheet.id for sheet in sheets if sheet.name == "REPLACE_WITH_SHEET_NAME"), None)

if sheet_id:
    # Load the sheet and map its columns for easier processing.
    sheet = smartsheet_client.Sheets.get_sheet(sheet_id)
    col_id_to_title = {col.id: col.title for col in sheet.columns}
    smartsheet_data = {}

    # Process each row in the sheet.
    for row in sheet.rows:
        row_id = format_value(row.id)
        row_data = {
            row_id_field: row_id,
            **{field_mapping.get(col_id_to_title[cell.column_id], col_id_to_title[cell.column_id]): format_date(cell.value) if isinstance(cell.value, str) and 'T' in cell.value else str(cell.value)
               for cell in row.cells if col_id_to_title.get(cell.column_id) in field_mapping}
        }
        # Store valid Parcel IDs in the list.
        parcel_id = row_data.get('Parcel_ID_Numbers')
        if parcel_id and parcel_id not in ['N/A', '', 'None']:
            parcel_id_list.append(parcel_id)
        smartsheet_data[row_id] = row_data

    # Append data to the feature service.
    append_to_feature_service(feature_service_url, smartsheet_data, field_mapping, parcel_id_list)

# Function to download attachments from Smartsheet rows.
def download_attachment(attachment, row_folder, sheet_id):
    """Download a single attachment to a local folder."""
    file_name = attachment.name
    local_file_path = os.path.join(row_folder, file_name)

    # Skip downloading if the file already exists locally.
    if os.path.exists(local_file_path):
        return

    # Get the download URL and save the file.
    attachment_details = smartsheet_client.Attachments.get_attachment(sheet_id, attachment.id)
    download_url = attachment_details.url
    response = requests.get(download_url, stream=True)
    if response.status_code == 200:
        with open(local_file_path, 'wb') as file:
            for chunk in response.iter_content(1024):
                file.write(chunk)
        print(f"Attachment '{file_name}' downloaded to: {local_file_path}")

# Use a ThreadPoolExecutor to handle multiple attachment downloads simultaneously.
with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
    future_to_attachment = {}

    for row in sheet.rows:
        row_id = row.id
        row_folder = os.path.join(download_folder, f"Row_{row_id}")
        os.makedirs(row_folder, exist_ok=True)

        # Retrieve attachments for this row and submit them for download.
        attachments = smartsheet_client.Attachments.list_row_attachments(sheet_id, row.id).data
        for attachment in attachments:
            future = executor.submit(download_attachment, attachment, row_folder, sheet_id)
            future_to_attachment[future] = attachment

    # Handle results and catch any errors during the download process.
    for future in concurrent.futures.as_completed(future_to_attachment):
        attachment = future_to_attachment[future]
        try:
            future.result()
        except Exception as exc:
            print(f"Error downloading attachment '{attachment.name}': {exc}")

# Initialize a log file to keep track of uploaded attachments.
log_file = os.path.join(download_folder, 'uploaded_attachments_log.xlsx')

# Create the log file if it doesn't exist.
if not os.path.exists(log_file):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Log"
    sheet.append(['Row_ID', 'Attachment'])
    workbook.save(log_file)

# Check if an attachment has already been uploaded.
def is_attachment_uploaded(row_id, file_name):
    """Check the log to see if this attachment has already been uploaded."""
    workbook = load_workbook(log_file)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if row == (str(row_id), file_name):
            return True
    return False

# Log a newly uploaded attachment.
def log_uploaded_attachment(row_id, file_name):
    """Add a record of the uploaded attachment to the log file."""
    workbook = load_workbook(log_file)
    sheet = workbook.active
    sheet.append([str(row_id), file_name])
    workbook.save(log_file)

# Prepare an attachment match table for uploading to the feature service.
match_table = os.path.join(arcpy.env.scratchGDB, "AttachmentMatchTable")
if arcpy.Exists(match_table):
    arcpy.Delete_management(match_table)

arcpy.CreateTable_management(arcpy.env.scratchGDB, "AttachmentMatchTable")
arcpy.AddField_management(match_table, "Row_ID", "TEXT")
arcpy.AddField_management(match_table, "ATTACHMENT", "TEXT")

# Populate the match table with attachments that need to be uploaded.
with arcpy.da.InsertCursor(match_table, ["Row_ID", "ATTACHMENT"]) as cursor:
    with arcpy.da.SearchCursor(feature_service_url, ["Row_ID"]) as search_cursor:
        for row in search_cursor:
            row_id = str(row[0])
            attachment_folder = os.path.join(download_folder, f'Row_{row_id}')

            if os.path.exists(attachment_folder):
                for file in os.listdir(attachment_folder):
                    file_path = os.path.join(attachment_folder, file)

                    # Only include files that haven't been logged as uploaded.
                    if not is_attachment_uploaded(str(row_id), file):
                        cursor.insertRow([row_id, file_path])
                        log_uploaded_attachment(str(row_id), file)

# Upload attachments to the feature service using the match table.
arcpy.management.AddAttachments(
    feature_service_url,
    "Row_ID",
    match_table,
    "Row_ID",
    "ATTACHMENT"
)

print("Attachments added successfully.")
